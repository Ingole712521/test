"""
Read companies from an Excel sheet, scrape the web for HR / careers / contact / info
emails, and write results back to the same file.

Looks for public addresses on company sites, career pages, and search snippets
(info@, hr@, careers@, contact@, recruitment@, etc.).

Usage:
  pip install -r requirements.txt
  python scrape_emails_from_excel.py --excel Company_email.xlsx
  python scrape_emails_from_excel.py --excel contacts.xlsx --only-missing
  python scrape_emails_from_excel.py --excel Company_email.xlsx --limit 20 --delay 2

Then send mail:
  python send_mail_merge.py --excel Company_email.xlsx --template email_template.txt
"""

from __future__ import annotations

import argparse
import re
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from job_hunter.company_lookup import find_emails_for_company

EMAIL_ALIASES = {
    "email",
    "e-mail",
    "emailid",
    "email id",
    "email_id",
    "mail",
    "recipient",
    "to",
    "email address",
}
COMPANY_ALIASES = {
    "company",
    "company name",
    "comapany name",
    "organization",
    "business",
    "firm",
    "client",
}
WEBSITE_ALIASES = {
    "website",
    "web site",
    "url",
    "company website",
    "site",
    "domain",
}

_SIMPLE_EMAIL = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
_NO_EMAIL_MARKERS = ("no email", "n/a", "na", "-", "none", "tbd", "not found")
_NUMBERED_COMPANY = re.compile(r"^\s*\d{1,4}\s*[\.\)]\s*")


def normalize_header(s: str) -> str:
    return str(s).strip().lower()


def pick_column(df: pd.DataFrame, aliases: set[str]) -> str | None:
    mapping = {normalize_header(c): c for c in df.columns}
    for alias in aliases:
        if alias in mapping:
            return mapping[alias]
    return None


def is_valid_email(raw: str) -> bool:
    s = str(raw).strip()
    if not s or s.lower() == "nan":
        return False
    low = s.lower()
    if any(m in low for m in _NO_EMAIL_MARKERS):
        return False
    if "@" not in s:
        return False
    first = re.split(r"[,;\n]+", s, maxsplit=1)[0].strip()
    return bool(_SIMPLE_EMAIL.match(first))


def clean_company_name(raw: str) -> str:
    s = str(raw).strip()
    if not s or s.lower() == "nan":
        return ""
    s = _NUMBERED_COMPANY.sub("", s).strip()
    return s


def is_valid_company(raw: str) -> bool:
    name = clean_company_name(raw)
    if not name:
        return False
    if name.isdigit():
        return False
    if len(name) < 2:
        return False
    return True


def load_company_only_sheet(path: Path) -> pd.DataFrame:
    """Single-column list (e.g. company_list.xlsx) → Company Name + empty Email ID."""
    raw = pd.read_excel(path, header=None)
    if raw.shape[1] < 1:
        raise SystemExit("Excel file has no columns.")

    col = raw.iloc[:, 0]
    names: list[str] = []
    for val in col:
        if not is_valid_company(str(val)):
            continue
        names.append(clean_company_name(str(val)))

    if not names:
        raise SystemExit("No company names found in the first column.")

    return pd.DataFrame({"Company Name": names, "Email ID": ""})


def load_sheet(path: Path, header_row: int | None) -> tuple[pd.DataFrame, int]:
    if header_row is not None:
        df = pd.read_excel(path, header=header_row)
        company_col = pick_column(df, COMPANY_ALIASES)
        email_col = pick_column(df, EMAIL_ALIASES)
        if company_col and email_col:
            return df, header_row
        if company_col and not email_col:
            df = df.copy()
            df["Email ID"] = ""
            return df, header_row
        raise SystemExit(
            f"--header-row {header_row} did not yield a company column. "
            f"Found: {list(df.columns)}"
        )

    for header in range(0, 8):
        df = pd.read_excel(path, header=header)
        company_col = pick_column(df, COMPANY_ALIASES)
        email_col = pick_column(df, EMAIL_ALIASES)
        if company_col and email_col:
            return df, header
        if company_col and not email_col:
            df = df.copy()
            df["Email ID"] = ""
            return df, header

    # No header row: try first column as company names only
    try:
        return load_company_only_sheet(path), 0
    except SystemExit:
        pass

    sample = pd.read_excel(path, header=0)
    raise SystemExit(
        "Could not find company names. Headers seen: "
        f"{list(sample.columns)}. For a one-column company list, put names in column A."
    )


def format_email_cell(emails: list[str]) -> str:
    if not emails:
        return "NO email id"
    return ", ".join(emails)


def save_dataframe(df: pd.DataFrame, path: Path, header_row: int) -> None:
    """Write dataframe back, preserving extra columns (e.g. Role, Status)."""
    out = path.with_suffix(".tmp.xlsx")
    df.to_excel(out, index=False, header=True)
    # If original had a title row above headers, re-insert via openpyxl copy is heavy;
    # most sheets use row 1 as headers (header_row 0).
    if header_row == 0:
        out.replace(path)
        return

    wb_new = load_workbook(out)
    ws_new = wb_new.active
    wb_old = load_workbook(path)
    ws_old = wb_old.active

    # Clear old data rows below header, copy from new sheet
    start_row = header_row + 2  # 1-based Excel row after header
    max_old = ws_old.max_row
    for r in range(start_row, max_old + 1):
        for c in range(1, ws_old.max_column + 1):
            ws_old.cell(row=r, column=c, value=None)

    for r_idx, row in enumerate(ws_new.iter_rows(min_row=2, values_only=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            if c_idx <= ws_old.max_column:
                ws_old.cell(row=r_idx, column=c_idx, value=val)

    wb_old.save(path)
    out.unlink(missing_ok=True)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Scrape HR/careers/contact/info emails for companies in an Excel sheet"
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=Path("company_list.xlsx"),
        help="Input Excel path (default: company_list.xlsx)",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=None,
        help="0-based row index for column headers (default: auto-detect)",
    )
    parser.add_argument(
        "--only-missing",
        action="store_true",
        help="Only look up rows without a valid email (default: re-scrape all)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        metavar="N",
        help="Max companies to look up (0 = all rows)",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=1.5,
        help="Seconds between company lookups (default: 1.5)",
    )
    parser.add_argument(
        "--max-emails",
        type=int,
        default=3,
        help="Max emails to store per company, comma-separated (default: 3)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("Company_email.xlsx"),
        help="Output Excel path (default: Company_email.xlsx; use same as --excel to overwrite)",
    )
    args = parser.parse_args()

    excel_in = args.excel.resolve()
    if not excel_in.is_file():
        raise SystemExit(f"Excel file not found: {excel_in}")

    excel_out = args.output.resolve()
    df, header_used = load_sheet(excel_in, args.header_row)

    company_col = pick_column(df, COMPANY_ALIASES)
    email_col = pick_column(df, EMAIL_ALIASES)
    website_col = pick_column(df, WEBSITE_ALIASES)
    if not company_col or not email_col:
        raise SystemExit(f"Missing columns. Found: {list(df.columns)}")

    print(f"Sheet: {excel_in.name}")
    print(f"Rows: {len(df)} | Company column: {company_col!r} | Email column: {email_col!r}")
    if website_col:
        print(f"Website column: {website_col!r}")
    print()

    looked_up = 0
    found_count = 0
    limit = args.limit if args.limit > 0 else len(df)

    for idx, row in df.iterrows():
        if looked_up >= limit:
            break

        company = clean_company_name(str(row[company_col]))
        if not is_valid_company(company):
            continue

        current = str(row[email_col]).strip() if pd.notna(row[email_col]) else ""
        if args.only_missing and is_valid_email(current):
            continue

        website = ""
        if website_col and pd.notna(row.get(website_col)):
            website = str(row[website_col]).strip()

        looked_up += 1
        print(f"[{looked_up}/{limit}] {company}")

        emails, source_url = find_emails_for_company(
            company,
            website=website,
            max_emails=args.max_emails,
        )
        cell = format_email_cell(emails)
        df.at[idx, email_col] = cell

        if emails:
            found_count += 1
            print(f"    -> {cell}")
            if source_url:
                print(f"    source: {source_url}")
        else:
            print("    -> no email found")

        save_dataframe(df, excel_out, header_used)

        if args.delay > 0 and looked_up < limit:
            time.sleep(args.delay)

    if looked_up == 0:
        print("No rows to process (all have emails? use without --only-missing).")
        return

    print(f"\nSaved to {excel_out}")
    print(f"Companies with email found this run: {found_count}/{looked_up}")


if __name__ == "__main__":
    main()
